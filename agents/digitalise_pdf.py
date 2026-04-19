"""
===============================================================================
DOCUMENT DIGITIZER WITH VALIDATION & HUMAN REVIEW
===============================================================================

WORKFLOW:
1. Extract Document → 2. LLM Processing → 3. Validation → 4. Decision:
                                                            ├─ Valid → Excel (Validated Sheet)
                                                            └─ Invalid → Human Review (Needs Review Sheet)

BENEFITS:
- Quality gate prevents bad data
- Human review for edge cases
- Maintains data integrity

===============================================================================
"""

from openai import OpenAI
import os
from dotenv import load_dotenv
import json
import pandas as pd
from datetime import datetime
import pdfplumber

load_dotenv()
client = OpenAI()

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using pdfplumber (handles tables well)."""
    text_parts = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            # Extract text
            page_text = page.extract_text()
            if page_text:
                text_parts.append(f"--- Page {page_num} ---")
                text_parts.append(page_text)
            
            # Extract tables as formatted text
            tables = page.extract_tables()
            for table_num, table in enumerate(tables, 1):
                text_parts.append(f"\n[Table {table_num}]")
                for row in table:
                    text_parts.append(" | ".join(str(cell) if cell else "" for cell in row))
    
    return "\n".join(text_parts)

# ============================================================================
# VALIDATION FUNCTION
# ============================================================================

def validate_invoice_data(invoice_data):
    """
    Validate extracted invoice data for quality and completeness.
    
    Args:
        invoice_data (dict): Extracted invoice data
    
    Returns:
        tuple: (is_valid: bool, errors: list of str)
    """
    errors = []
    
    # Skip validation if extraction failed
    if "error" in invoice_data:
        return False, ["LLM extraction failed"]
    
    # 1. REQUIRED FIELDS CHECK
    required_fields = ["invoice_number", "customer_name", "total_amount"]
    for field in required_fields:
        value = invoice_data.get(field)
        if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
            errors.append(f"Missing required field: {field}")
    
    # 2. DATA TYPE VALIDATION
    numeric_fields = ["subtotal", "cgst", "sgst", "total_amount"]
    for field in numeric_fields:
        value = invoice_data.get(field)
        if value is not None:
            if not isinstance(value, (int, float)):
                errors.append(f"{field} must be numeric, got: {type(value).__name__}")
            elif value < 0:
                errors.append(f"{field} cannot be negative: {value}")
    
    # 3. GST FORMAT VALIDATION (Indian GST: 15 alphanumeric)
    for gst_field in ["company_gst", "customer_gst"]:
        gst_value = invoice_data.get(gst_field)
        if gst_value and gst_value != "null":
            gst_str = str(gst_value).strip()
            if len(gst_str) != 15:
                errors.append(f"{gst_field} must be 15 characters: {gst_str}")
            elif not gst_str.isalnum():
                errors.append(f"{gst_field} must be alphanumeric: {gst_str}")
    
    # 4. MATHEMATICAL CONSISTENCY
    subtotal = invoice_data.get("subtotal")
    cgst = invoice_data.get("cgst")
    sgst = invoice_data.get("sgst")
    total = invoice_data.get("total_amount")
    
    if all(isinstance(v, (int, float)) for v in [subtotal, cgst, sgst, total]):
        calculated_total = subtotal + cgst + sgst
        # Allow small floating point differences
        if abs(calculated_total - total) > 0.01:
            errors.append(
                f"Math error: Subtotal({subtotal}) + CGST({cgst}) + SGST({sgst}) "
                f"= {calculated_total}, but Total = {total}"
            )
    
    # 5. DATE VALIDATION (invoice_date should not be in future)
    invoice_date_str = invoice_data.get("invoice_date")
    if invoice_date_str:
        try:
            # Try common date formats
            from dateutil import parser
            invoice_date = parser.parse(str(invoice_date_str))
            current_date = datetime.now()
            
            # Check if date is in future
            if invoice_date > current_date:
                errors.append(f"Invoice date is in future: {invoice_date_str}")
        except:
            # If date parsing fails, add as warning but don't fail validation
            errors.append(f"Could not parse invoice_date format: {invoice_date_str}")
    
    # 6. BUSINESS LOGIC - Total amount should be positive
    total_amt = invoice_data.get("total_amount")
    if isinstance(total_amt, (int, float)) and total_amt <= 0:
        errors.append(f"Total amount must be positive: {total_amt}")
    
    # Return validation result
    is_valid = len(errors) == 0
    return is_valid, errors

# ============================================================================
# EXTRACTION FUNCTION - INVOICE VERSION
# ============================================================================

def extract_invoice_data(pdf_path):
    """
    Extract data from invoice PDF document.
    
    Args:
        pdf_path (str): Path to PDF file
    
    Returns:
        dict: Extracted invoice data exactly as shown in document
    """
    
    prompt = """
Extract ALL information from this invoice document.

Return ONLY valid JSON with these exact fields (use null if not found):

{
  "company_name": "text",
  "company_gst": "text",
  "invoice_number": "text",
  "invoice_date": "text",
  "po_number": "text",
  "customer_name": "text",
  "customer_gst": "text",
  "subtotal": number,
  "cgst": number,
  "sgst": number,
  "total_amount": number,
  "extraction_timestamp": "ISO timestamp"
}

IMPORTANT: 
- Extract text EXACTLY as shown 
- Convert amounts to numbers (remove ₹, Rs, $, commas)
- Use null for missing fields
- Return only JSON, no extra text
"""
    
    # Extract text from PDF
    pdf_text = extract_text_from_pdf(pdf_path)
    
    # Get response from LLM
    messages = [{"role": "user", "content": prompt + f"\n\nDocument:\n{pdf_text}"}]
    api_response = client.responses.create(
        model="gpt-4o-mini",
        input=messages,
        temperature=0
    )
    response = api_response.output_text
    
    # Parse JSON
    try:
        cleaned = response.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        
        result = json.loads(cleaned.strip())
        return result
    except json.JSONDecodeError as e:
        print(f"⚠️  JSON Error: {e}")
        return {"error": "Parse failed", "raw_response": response}

# ============================================================================
# EXCEL EXPORT - DUAL SHEET (VALIDATED + NEEDS REVIEW)
# ============================================================================

def export_to_excel_with_validation(validated_invoices, review_invoices, output_file="invoices_extracted.xlsx"):
    """
    Export data to Excel with TWO sheets:
    - Sheet 1: Validated_Invoices (passed validation)
    - Sheet 2: Needs_Review (failed validation)
    
    Args:
        validated_invoices (list): List of validated invoice dicts
        review_invoices (list): List of invoices needing review with errors
        output_file (str): Output file path
    """
    # Column mapping for better readability
    column_mapping = {
        "company_name": "Company Name",
        "company_gst": "Company GST",
        "invoice_number": "Invoice No",
        "invoice_date": "Date",
        "po_number": "PO Number",
        "customer_name": "Customer",
        "customer_gst": "Customer GST",
        "subtotal": "Subtotal",
        "cgst": "CGST",
        "sgst": "SGST",
        "total_amount": "Total Amount",
        "extraction_timestamp": "Extracted On",
        "validation_errors": "Validation Errors"
    }
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Prepare dataframes
    validated_df = pd.DataFrame(validated_invoices) if validated_invoices else pd.DataFrame()
    review_df = pd.DataFrame(review_invoices) if review_invoices else pd.DataFrame()
    
    # Rename columns
    if not validated_df.empty:
        validated_df = validated_df.rename(columns=column_mapping)
    if not review_df.empty:
        review_df = review_df.rename(columns=column_mapping)
    
    # Handle existing file - append to appropriate sheets
    if os.path.exists(output_file):
        print(f"📄 Found existing file: {output_file}")
        try:
            existing_validated = pd.read_excel(output_file, sheet_name='Validated_Invoices')
            existing_review = pd.read_excel(output_file, sheet_name='Needs_Review')
            
            if not validated_df.empty:
                validated_df = pd.concat([existing_validated, validated_df], ignore_index=True)
            else:
                validated_df = existing_validated
                
            if not review_df.empty:
                review_df = pd.concat([existing_review, review_df], ignore_index=True)
            else:
                review_df = existing_review
                
            print("✅ Appended to existing sheets")
        except Exception as e:
            print(f"⚠️  Could not read existing sheets: {e}")
            print("   Creating new sheets...")
    
    # Write to Excel with both sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: Validated Invoices
        if not validated_df.empty:
            validated_df.to_excel(writer, sheet_name='Validated_Invoices', index=False)
            ws_validated = writer.sheets['Validated_Invoices']
            
            # Format header - Green theme
            for cell in ws_validated[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust columns
            for column in ws_validated.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_validated.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        # Sheet 2: Needs Review
        if not review_df.empty:
            review_df.to_excel(writer, sheet_name='Needs_Review', index=False)
            ws_review = writer.sheets['Needs_Review']
            
            # Format header - Red theme
            for cell in ws_review[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust columns
            for column in ws_review.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_review.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Highlight validation errors column
            error_col_idx = list(review_df.columns).index('Validation Errors') + 1
            for row in range(2, ws_review.max_row + 1):
                cell = ws_review.cell(row=row, column=error_col_idx)
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    print(f"\n✅ Excel saved: {output_file}")
    print(f"   ✓ Validated: {len(validated_invoices)} invoices")
    print(f"   ⚠ Needs Review: {len(review_invoices)} invoices")

# ============================================================================
# MAIN TEST
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*80)
    print("INVOICE DIGITIZER WITH VALIDATION & HUMAN REVIEW")
    print("="*80)
    
    # Check for PDF
    pdf_file = "sample_invoice.pdf"
    
    if not os.path.exists(pdf_file):
        print("\n⚠️  PDF not found!")
        print(f"   Place '{pdf_file}' in this directory")
        exit(1)
    
    # STEP 1: Extract data from PDF
    print(f"\n[STEP 1] 📄 Extracting from PDF: {pdf_file}")
    extracted = extract_invoice_data(pdf_path=pdf_file)
    
    print("\n📊 EXTRACTED DATA:")
    print(json.dumps(extracted, indent=2))
    
    # STEP 2: Validate extracted data
    print("\n[STEP 2] 🔍 Validating data quality...")
    is_valid, validation_errors = validate_invoice_data(extracted)
    
    if is_valid:
        print("   ✅ PASSED - Data is valid!")
    else:
        print(f"   ⚠️  FAILED - Found {len(validation_errors)} validation error(s):")
        for i, error in enumerate(validation_errors, 1):
            print(f"      {i}. {error}")
    
    # STEP 3: Route to appropriate destination
    print("\n[STEP 3] 📊 Routing data...")
    
    validated_invoices = []
    review_invoices = []
    
    if is_valid:
        validated_invoices.append(extracted)
        print("   → Sending to: Validated_Invoices sheet")
    else:
        # Add validation errors to the invoice data
        extracted["validation_errors"] = " | ".join(validation_errors)
        review_invoices.append(extracted)
        print("   → Sending to: Needs_Review sheet (Human review required)")
    
    # STEP 4: Export to Excel
    print("\n[STEP 4] 📤 Exporting to Excel...")
    export_to_excel_with_validation(
        validated_invoices=validated_invoices,
        review_invoices=review_invoices,
        output_file="invoices_extracted.xlsx"
    )
    
    print("\n" + "="*80)
    print("✅ COMPLETE! Check 'invoices_extracted.xlsx'")
    print("   - Green sheet = Validated (ready to use)")
    print("   - Red sheet = Needs Review (check validation errors)")
    print("="*80)
