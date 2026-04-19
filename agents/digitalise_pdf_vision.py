"""
===============================================================================
ADVANCED: DOCUMENT DIGITIZER - VISION MODEL FOR COMPLEX PDFs
===============================================================================

BEST FOR:
- Complex tabular structures
- Scanned documents (image-based PDFs)
- Multi-column layouts
- Handwritten annotations
- Forms with checkboxes/stamps
- Documents where text extraction fails

WORKFLOW:
1. PDF → High-Quality Image (PyMuPDF)
2. Image → Vision LLM (GPT-4o) → Structured Data
3. Validation (6 quality checks)
4. Decision: Valid → Excel Green | Invalid → Excel Red (Human Review)

===============================================================================
"""

from openai import OpenAI
import os
from dotenv import load_dotenv
import json
import pandas as pd
from datetime import datetime
import fitz  # PyMuPDF
from PIL import Image
import io
import base64

load_dotenv()
client = OpenAI()

# ============================================================================
# PDF TO IMAGE CONVERSION
# ============================================================================

def pdf_to_base64_images(pdf_path, dpi=300):
    """
    Convert PDF pages to base64-encoded images for vision API.
    
    Args:
        pdf_path (str): Path to PDF file
        dpi (int): Resolution (300 recommended for documents)
    
    Returns:
        list: Base64-encoded images (one per page)
    """
    print(f"  📸 Converting PDF to images (DPI: {dpi})...")
    
    pdf = fitz.open(pdf_path)
    images_base64 = []
    
    for page_num in range(len(pdf)):
        page = pdf[page_num]
        
        # Calculate zoom factor for desired DPI
        zoom = dpi / 72  # 72 is default DPI
        mat = fitz.Matrix(zoom, zoom)
        
        # Render page to high-quality image
        pix = page.get_pixmap(matrix=mat)
        
        # Convert to PIL Image
        img = Image.open(io.BytesIO(pix.tobytes()))
        
        # Convert to base64 for API
        buffered = io.BytesIO()
        img.save(buffered, format='PNG')
        img_base64 = base64.b64encode(buffered.getvalue()).decode()
        
        images_base64.append(img_base64)
        print(f"     ✓ Page {page_num + 1} rendered")
    
    pdf.close()
    print(f"  ✅ Converted {len(images_base64)} page(s) to images\n")
    return images_base64

# ============================================================================
# VISION MODEL EXTRACTION
# ============================================================================

def extract_invoice_data_with_vision(pdf_path):
    """
    Extract invoice data using Vision Model (handles complex PDFs).
    
    Args:
        pdf_path (str): Path to PDF file
    
    Returns:
        dict: Extracted invoice data
    """
    
    prompt = """
You are analyzing an invoice document image. Extract ALL invoice information.

Return ONLY valid JSON with these exact fields (use null if not found):

{
  "company_name": "text",
  "company_gst": "text",
  "invoice_number": "text",
  "invoice_date": "text",
  "po_number": "text",
  "customer_name": "text",
  "customer_gst": "text",
  "subtotal": number,
  "cgst": number,
  "sgst": number,
  "total_amount": number,
  "extraction_timestamp": "ISO timestamp"
}

IMPORTANT INSTRUCTIONS:
- Extract text EXACTLY as shown in the image
- Pay attention to table structures and layouts
- Convert amounts to numbers (remove currency symbols, commas)
- For GST/Tax numbers, preserve exact format
- Use null for missing/unclear fields
- Return ONLY the JSON, no explanations

If you see merged cells or complex tables, carefully identify which values belong to which fields.
"""
    
    # Step 1: Convert PDF to images
    images_base64 = pdf_to_base64_images(pdf_path, dpi=300)
    
    # Step 2: Process first page with vision model
    # (For multi-page invoices, you can loop through all pages)
    print("  🤖 Processing with Vision Model (GPT-4o)...")
    
    try:
        response = client.chat.completions.create(
            model="gpt-4o",  # Vision-enabled model
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": prompt
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{images_base64[0]}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=1000,
            temperature=0
        )
        
        result_text = response.choices[0].message.content
        print("  ✅ Vision model processing complete\n")
        
        # Parse JSON response
        try:
            # Clean markdown code blocks if present
            cleaned = result_text.strip()
            if cleaned.startswith("```json"):
                cleaned = cleaned[7:]
            if cleaned.startswith("```"):
                cleaned = cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            
            result = json.loads(cleaned.strip())
            result["extraction_timestamp"] = datetime.now().isoformat()
            return result
            
        except json.JSONDecodeError as e:
            print(f"  ⚠️  JSON Parse Error: {e}")
            return {"error": "JSON parse failed", "raw_response": result_text}
    
    except Exception as e:
        print(f"  ❌ Vision API Error: {e}")
        return {"error": str(e)}

# ============================================================================
# VALIDATION FUNCTION
# ============================================================================

def validate_invoice_data(invoice_data):
    """
    Validate extracted invoice data for quality and completeness.
    
    Args:
        invoice_data (dict): Extracted invoice data
    
    Returns:
        tuple: (is_valid: bool, errors: list of str)
    """
    errors = []
    
    # Skip validation if extraction failed
    if "error" in invoice_data:
        return False, [f"Extraction failed: {invoice_data.get('error', 'Unknown error')}"]
    
    # 1. REQUIRED FIELDS CHECK
    required_fields = ["invoice_number", "customer_name", "total_amount"]
    for field in required_fields:
        value = invoice_data.get(field)
        if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
            errors.append(f"Missing required field: {field}")
    
    # 2. DATA TYPE VALIDATION
    numeric_fields = ["subtotal", "cgst", "sgst", "total_amount"]
    for field in numeric_fields:
        value = invoice_data.get(field)
        if value is not None:
            if not isinstance(value, (int, float)):
                errors.append(f"{field} must be numeric, got: {type(value).__name__}")
            elif value < 0:
                errors.append(f"{field} cannot be negative: {value}")
    
    # 3. GST FORMAT VALIDATION (Indian GST: 15 alphanumeric)
    for gst_field in ["company_gst", "customer_gst"]:
        gst_value = invoice_data.get(gst_field)
        if gst_value and gst_value != "null":
            gst_str = str(gst_value).strip()
            if len(gst_str) != 15:
                errors.append(f"{gst_field} must be 15 characters: {gst_str}")
            elif not gst_str.isalnum():
                errors.append(f"{gst_field} must be alphanumeric: {gst_str}")
    
    # 4. MATHEMATICAL CONSISTENCY
    subtotal = invoice_data.get("subtotal")
    cgst = invoice_data.get("cgst")
    sgst = invoice_data.get("sgst")
    total = invoice_data.get("total_amount")
    
    if all(isinstance(v, (int, float)) for v in [subtotal, cgst, sgst, total]):
        calculated_total = subtotal + cgst + sgst
        # Allow small floating point differences
        if abs(calculated_total - total) > 0.01:
            errors.append(
                f"Math error: Subtotal({subtotal}) + CGST({cgst}) + SGST({sgst}) "
                f"= {calculated_total}, but Total = {total}"
            )
    
    # 5. DATE VALIDATION
    invoice_date_str = invoice_data.get("invoice_date")
    if invoice_date_str:
        try:
            from dateutil import parser
            invoice_date = parser.parse(str(invoice_date_str))
            current_date = datetime.now()
            
            if invoice_date > current_date:
                errors.append(f"Invoice date is in future: {invoice_date_str}")
        except:
            errors.append(f"Could not parse invoice_date format: {invoice_date_str}")
    
    # 6. BUSINESS LOGIC
    total_amt = invoice_data.get("total_amount")
    if isinstance(total_amt, (int, float)) and total_amt <= 0:
        errors.append(f"Total amount must be positive: {total_amt}")
    
    # Return validation result
    is_valid = len(errors) == 0
    return is_valid, errors

# ============================================================================
# EXCEL EXPORT - DUAL SHEET (VALIDATED + NEEDS REVIEW)
# ============================================================================

def export_to_excel_with_validation(validated_invoices, review_invoices, output_file="invoices_vision.xlsx"):
    """
    Export data to Excel with TWO sheets:
    - Sheet 1: Validated_Invoices (passed validation)
    - Sheet 2: Needs_Review (failed validation)
    
    Args:
        validated_invoices (list): List of validated invoice dicts
        review_invoices (list): List of invoices needing review with errors
        output_file (str): Output file path
    """
    
    # Column mapping
    column_mapping = {
        "company_name": "Company Name",
        "company_gst": "Company GST",
        "invoice_number": "Invoice No",
        "invoice_date": "Date",
        "po_number": "PO Number",
        "customer_name": "Customer",
        "customer_gst": "Customer GST",
        "subtotal": "Subtotal",
        "cgst": "CGST",
        "sgst": "SGST",
        "total_amount": "Total Amount",
        "extraction_timestamp": "Extracted On",
        "validation_errors": "Validation Errors"
    }
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Prepare dataframes
    validated_df = pd.DataFrame(validated_invoices) if validated_invoices else pd.DataFrame()
    review_df = pd.DataFrame(review_invoices) if review_invoices else pd.DataFrame()
    
    # Rename columns
    if not validated_df.empty:
        validated_df = validated_df.rename(columns=column_mapping)
    if not review_df.empty:
        review_df = review_df.rename(columns=column_mapping)
    
    # Handle existing file
    if os.path.exists(output_file):
        print(f"📄 Found existing file: {output_file}")
        try:
            existing_validated = pd.read_excel(output_file, sheet_name='Validated_Invoices')
            existing_review = pd.read_excel(output_file, sheet_name='Needs_Review')
            
            if not validated_df.empty:
                validated_df = pd.concat([existing_validated, validated_df], ignore_index=True)
            else:
                validated_df = existing_validated
                
            if not review_df.empty:
                review_df = pd.concat([existing_review, review_df], ignore_index=True)
            else:
                review_df = existing_review
                
            print("✅ Appended to existing sheets")
        except Exception as e:
            print(f"⚠️  Could not read existing sheets: {e}")
            print("   Creating new sheets...")
    
    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: Validated Invoices
        if not validated_df.empty:
            validated_df.to_excel(writer, sheet_name='Validated_Invoices', index=False)
            ws_validated = writer.sheets['Validated_Invoices']
            
            # Format header - Green
            for cell in ws_validated[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust columns
            for column in ws_validated.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_validated.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        # Sheet 2: Needs Review
        if not review_df.empty:
            review_df.to_excel(writer, sheet_name='Needs_Review', index=False)
            ws_review = writer.sheets['Needs_Review']
            
            # Format header - Red
            for cell in ws_review[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust columns
            for column in ws_review.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_review.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Highlight validation errors
            if 'Validation Errors' in review_df.columns:
                error_col_idx = list(review_df.columns).index('Validation Errors') + 1
                for row in range(2, ws_review.max_row + 1):
                    cell = ws_review.cell(row=row, column=error_col_idx)
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    print(f"\n✅ Excel saved: {output_file}")
    print(f"   ✓ Validated: {len(validated_invoices)} invoices")
    print(f"   ⚠ Needs Review: {len(review_invoices)} invoices")

# ============================================================================
# MAIN TEST
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*80)
    print("VISION MODEL INVOICE DIGITIZER - FOR COMPLEX PDFs")
    print("="*80)
    print("📌 Use this version for:")
    print("   • Complex tables & layouts")
    print("   • Scanned documents")
    print("   • Handwritten annotations")
    print("   • Image-based PDFs")
    print("="*80)
    
    # Check for PDF
    pdf_file = "sample_invoice.pdf"
    
    if not os.path.exists(pdf_file):
        print("\n⚠️  PDF not found!")
        print(f"   Place '{pdf_file}' in this directory")
        exit(1)
    
    # STEP 1: Extract with Vision Model
    print(f"\n[STEP 1] 👁️  Extracting with Vision Model: {pdf_file}")
    extracted = extract_invoice_data_with_vision(pdf_path=pdf_file)
    
    print("📊 EXTRACTED DATA:")
    print(json.dumps(extracted, indent=2))
    
    # STEP 2: Validate
    print("\n[STEP 2] 🔍 Validating data quality...")
    is_valid, validation_errors = validate_invoice_data(extracted)
    
    if is_valid:
        print("   ✅ PASSED - Data is valid!")
    else:
        print(f"   ⚠️  FAILED - Found {len(validation_errors)} error(s):")
        for i, error in enumerate(validation_errors, 1):
            print(f"      {i}. {error}")
    
    # STEP 3: Route
    print("\n[STEP 3] 📊 Routing data...")
    
    validated_invoices = []
    review_invoices = []
    
    if is_valid:
        validated_invoices.append(extracted)
        print("   → Validated_Invoices sheet (Green)")
    else:
        extracted["validation_errors"] = " | ".join(validation_errors)
        review_invoices.append(extracted)
        print("   → Needs_Review sheet (Red) - Human review required")
    
    # STEP 4: Export
    print("\n[STEP 4] 📤 Exporting to Excel...")
    export_to_excel_with_validation(
        validated_invoices=validated_invoices,
        review_invoices=review_invoices,
        output_file="invoices_vision.xlsx"
    )
    
    print("\n" + "="*80)
    print("✅ COMPLETE! Check 'invoices_vision.xlsx'")
    print("   🟢 Green sheet = Validated (ready)")
    print("   🔴 Red sheet = Needs Review")
    print("="*80)
